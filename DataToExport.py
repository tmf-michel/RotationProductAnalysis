from SalesListFunctions import mostSelledProductsRaw, mostSelledProductsMonthlyRaw
from SearchListFunctions import mostSearchedProductsRaw
from SalesCategoryFunctions import lessSelledByCategoryRaw, lessSelledByMonthlyCategoryRaw
from SearchCategoryFunctions import lessSearchedByCategoryRaw
from Reviews import bestReviewsRaw, bestMonthlyReviewsRaw, worstReviewsRaw, worstMonthlyReviewsRaw, reviewsByCategoryRaw, monthlyReviewsByCategoryRaw
from TotalSales import salesPerMonthRaw
from openpyxl import Workbook
wb = Workbook()
ruta = 'REPORTE-01-CHAVARIN-JORGE.xlsx'
def exportData(qtyLow, qtyHigh, year):
  rawData = {}
  mostSelledProducts = mostSelledProductsRaw(qtyLow, year)
  # Por si queremos agregar encabezados a la tabla
  mostSelledProductsHeader =  [['Producto', 'Cantidad de Ventas']]
  mostSelledProducts = mostSelledProductsHeader + mostSelledProducts
  rawData.update({'Productos mas vendidos en el año':mostSelledProducts})

  hoja = wb.active
  hoja.title = "Ventas"
  fila = 1
  hoja.cell(column=1, row=fila, value='5 Productos más vendidos en el año')
  fila += 1

  for i in range(len(mostSelledProducts)):
    hoja.cell(column=1, row=fila, value=mostSelledProducts[i][0])
    hoja.cell(column=2, row=fila, value=mostSelledProducts[i][1])
    fila+=1
  fila+=1
##########################################################################################################
##########################################################################################################
  mostSelledProductsMonthly = mostSelledProductsMonthlyRaw(qtyLow, year)
  rawData.update({'Productos mas vendidos por mes': mostSelledProductsMonthly})

  hoja.cell(column=1, row=fila, value='5 Productos más vendidos por mes')
  fila += 1

  for month, data in mostSelledProductsMonthly.items():
    hoja.cell(column=1, row=fila, value="Mes")
    hoja.cell(column=2, row=fila, value='Producto')
    hoja.cell(column=3, row=fila, value='Ventas')
    fila+=1
    for i in data:
      hoja.cell(column=1, row=fila, value=month)
      hoja.cell(column=2, row=fila, value=data[0][0])
      hoja.cell(column=3, row=fila, value=data[1][1])
      fila+=1
  fila+=1
##########################################################################################################
##########################################################################################################
  mostSearchedProducts = mostSearchedProductsRaw(qtyHigh)
  # Encabezados de tabla
  # mostSearchedProductsHeaders  = [['Productos', 'Cantidad de Búsquedas']]
  # mostSearchedProducts = mostSearchedProductsHeaders + mostSearchedProducts
  rawData.update({'Productos mas buscados':mostSearchedProducts})

  hoja.cell(column=1, row=fila, value='10 Productos más buscados')
  fila += 1
  for i in range(len(mostSearchedProducts)):
    hoja.cell(column=1, row=fila, value=mostSearchedProducts[i][0])
    hoja.cell(column=2, row=fila, value=mostSearchedProducts[i][1])
    fila+=1
  fila+=1
##########################################################################################################
##########################################################################################################
  lessSelledByCategory = lessSelledByCategoryRaw(qtyLow, year)
  # print(lessSelledByCategory)
  # Encabezados de tabla
  # lessSearchedByCategory = {'Productos menos vendidos por categoria': lessSearchedByCategory}
  rawData.update({'Productos menos vendidos por categoria': lessSelledByCategory})

  hoja.cell(column=1, row=fila, value='5 Productos menos vendidos por categoria')
  fila += 1
  hoja.cell(column=1, row=fila, value="Categoría")
  hoja.cell(column=2, row=fila, value='Producto')
  hoja.cell(column=3, row=fila, value='Ventas')
  fila+=1
  for category, data in mostSelledProductsMonthly.items():
    for i in data:
      hoja.cell(column=1, row=fila, value=category)
      hoja.cell(column=2, row=fila, value=data[0][0])
      hoja.cell(column=3, row=fila, value=data[0][1])
      fila+=1
  fila+=1
##########################################################################################################
##########################################################################################################
  lessSelledByMonthlyCategory = lessSelledByMonthlyCategoryRaw(qtyLow, year)
  rawData.update({'Productos menos vendidos mensualemente por categoria': lessSelledByMonthlyCategory})

  hoja.cell(column=1, row=fila, value='5 Productos menos vendidos por categoria')
  fila += 1
  # print(lessSelledByMonthlyCategory)
  hoja.cell(column=1, row=fila, value="Mes")
  hoja.cell(column=2, row=fila, value="Categoria")
  hoja.cell(column=3, row=fila, value="ID")
  hoja.cell(column=4, row=fila, value="Nombre")
  hoja.cell(column=5, row=fila, value="Ventas")
  fila+=1
  for data in lessSelledByMonthlyCategory:
    for month in data:
      for category in data[month]:
        # print('cat', data[month][category])
        for product in data[month][category]:
          hoja.cell(column=1, row=fila, value=month)
          hoja.cell(column=2, row=fila, value=category)
          hoja.cell(column=3, row=fila, value=product[0])
          hoja.cell(column=4, row=fila, value=product[1])
          hoja.cell(column=5, row=fila, value=product[2])
          fila+=1

      
##########################################################################################################
##########################################################################################################
  lessSearchedByCategory = lessSearchedByCategoryRaw(qtyHigh)
  # print(lessSearchedByCategory)
  # Encabezados de tabla
  # lessSearchedByCategory = {'Productos menos buscados, por categoria': lessSearchedByCategory}
  rawData.update({'Productos menos buscados por categoria': lessSearchedByCategory})

  hoja.cell(column=1, row=fila, value='10 Productos menos buscados por categoria')
  fila += 1
  hoja.cell(column=1, row=fila, value="Categoria")
  hoja.cell(column=2, row=fila, value="ID")
  hoja.cell(column=3, row=fila, value="Nombre")
  hoja.cell(column=4, row=fila, value="Búsquedas")
  fila+=1
  for category, products in lessSearchedByCategory.items():
    for product in products:
      hoja.cell(column=1, row=fila, value=category)
      hoja.cell(column=2, row=fila, value=product[0])
      hoja.cell(column=3, row=fila, value=product[1])
      hoja.cell(column=4, row=fila, value=product[2])
      fila+=1
##########################################################################################################
##########################################################################################################
  bestReviews = bestReviewsRaw(qtyLow, year)
  # Encabezados de tabla
  # bestReviewsHeaders = [['Producto', 'Score Promedio']]
  # bestReviews = bestReviewsHeaders + bestReviews
  rawData.update({'Productos con mejores reseñas':bestReviews})
  
  hoja.cell(column=1, row=fila, value='5 Productos con mejores reseñas')
  fila += 1
  hoja.cell(column=1, row=fila, value="Producto")
  hoja.cell(column=2, row=fila, value="Score")
  fila+=1
  for product in bestReviews:
    hoja.cell(column=1, row=fila, value=product[0])
    hoja.cell(column=2, row=fila, value=product[1])
    fila+=1
##########################################################################################################
##########################################################################################################
  bestMonthlyReviews = bestMonthlyReviewsRaw(qtyLow, year)
  rawData.update({'Productos con mejores reseñas, por mes': bestMonthlyReviews})

  hoja.cell(column=1, row=fila, value='5 Productos con mejores reseñas por mes')
  fila += 1
  hoja.cell(column=1, row=fila, value="Mes")
  hoja.cell(column=2, row=fila, value="Producto")
  hoja.cell(column=3, row=fila, value="Score")
  fila+=1
  for month, products in bestMonthlyReviews.items():
    for product in products:
      hoja.cell(column=1, row=fila, value=month)
      hoja.cell(column=2, row=fila, value=product[0])
      hoja.cell(column=3, row=fila, value=product[1])
      fila+=1
##########################################################################################################
##########################################################################################################
  worstReviews = worstReviewsRaw(qtyLow, year)
  # Encabezados de tabla
  # worstReviewsHeaders = [['Producto', 'Score Promedio']]
  # worstReviews = worstReviewsHeaders + worstReviews
  rawData.update({'Productos con peores reseñas':worstReviews})

  hoja.cell(column=1, row=fila, value='5 Productos con peores reseñas')
  fila += 1
  hoja.cell(column=1, row=fila, value="Producto")
  hoja.cell(column=2, row=fila, value="Score")
  fila+=1
  for product in worstReviews:
    hoja.cell(column=1, row=fila, value=product[0])
    hoja.cell(column=2, row=fila, value=product[1])
    fila+=1
##########################################################################################################
##########################################################################################################
  worstMonthlyReviews = worstMonthlyReviewsRaw(qtyLow, year)
  rawData.update({'Productos con peores reseñas, por mes': worstMonthlyReviews})

  hoja.cell(column=1, row=fila, value='5 Productos con peores reseñas por mes')
  fila += 1
  hoja.cell(column=1, row=fila, value="Mes")
  hoja.cell(column=2, row=fila, value="Producto")
  hoja.cell(column=3, row=fila, value="Score")
  fila+=1
  for month, product in worstMonthlyReviews.items():
    for product in products:
      hoja.cell(column=1, row=fila, value=month)
      hoja.cell(column=2, row=fila, value=product[0])
      hoja.cell(column=3, row=fila, value=product[1])
      fila+=1
##########################################################################################################
##########################################################################################################
  reviewsByCategory = reviewsByCategoryRaw(year)
  # print(reviewsByCategory)
  # Encabezados de tabla
  # reviewsByCategoryHeader = [['Producto', 'Score']]
  # reviewsByCategory = reviewsByCategoryHeader + reviewsByCategory
  rawData.update({'Reseñas por categoría:': reviewsByCategory})
  hoja.cell(column=1, row=fila, value='Reseñas por categoria')
  fila += 1
  hoja.cell(column=1, row=fila, value="Categoria")
  hoja.cell(column=2, row=fila, value="Score")
  fila+=1
  for category in reviewsByCategory:
    hoja.cell(column=1, row=fila, value=category[0])
    hoja.cell(column=2, row=fila, value=category[1])
    fila+=1
##########################################################################################################
##########################################################################################################
  monthlyReviewsByCategory = monthlyReviewsByCategoryRaw(year)
  rawData.update({'Reseñas mensuales por cateoria: ': monthlyReviewsByCategory})
  
  hoja.cell(column=1, row=fila, value='Reseñas mensuales por categoria')
  fila += 1
  hoja.cell(column=1, row=fila, value="Categoria")
  hoja.cell(column=2, row=fila, value="Score")
  fila+=1
  months = [
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'
  ]
  cont = 0
  for month in monthlyReviewsByCategory:
    for category in month:
      hoja.cell(column=1, row=fila, value=months[cont])
      hoja.cell(column=2, row=fila, value=category[0])
      hoja.cell(column=3, row=fila, value=category[1])
      fila+=1
    cont+=1
##########################################################################################################
##########################################################################################################
  salesPerMonth = salesPerMonthRaw(year)
  # Encabezados de tabla
  # salesPerMonthHeaders = [['Fecha', 'Cantidad de Ventas', 'Ingresos']]
  # salesPerMonth = salesPerMonthHeaders + salesPerMonth
  rawData.update({'Ventas promedio mensuales': salesPerMonth})
  
  hoja.cell(column=1, row=fila, value='Ventas promedio mensuales')
  fila += 1
  hoja.cell(column=1, row=fila, value="Mes")
  hoja.cell(column=2, row=fila, value="Cantidad de Ventas")
  hoja.cell(column=3, row=fila, value="Ingresos")
  fila+=1
  for venta in salesPerMonth:
    hoja.cell(column=1, row=fila, value=venta[0])
    hoja.cell(column=2, row=fila, value=venta[1])
    hoja.cell(column=3, row=fila, value=venta[2])
    fila+=1
    
##########################################################################################################
##########################################################################################################

  for section in rawData:
    print('\n')
    print(section+':', rawData[section])
  print('\n')
  
  
  wb.save(filename = ruta)